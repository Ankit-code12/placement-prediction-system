import warnings
warnings.filterwarnings('ignore', category=UserWarning)
import os
import pickle
import numpy as np
import pandas as pd
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.models import User
from django.db.models import Count, Avg
import json
from datetime import datetime
import io

# Try to import models
try:
    from .models import Prediction
    MODELS_AVAILABLE = True
except ImportError:
    MODELS_AVAILABLE = False
    print("Models not yet migrated. Run: python manage.py makemigrations && python manage.py migrate")

# ReportLab for PDF
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib.colors import HexColor
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# OpenPyXL for Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

model_path = os.path.join(BASE_DIR, 'predictor', 'placement_model.pkl')
scaler_path = os.path.join(BASE_DIR, 'predictor', 'scaler.pkl')

model = pickle.load(open(model_path, 'rb'))
scaler = pickle.load(open(scaler_path, 'rb'))

# ==================== JOBS DATABASE ====================
JOBS_DATABASE = [
    {
        'title': 'Software Engineer',
        'company': 'Google',
        'min_cgpa': 8.0,
        'min_skill': 85,
        'salary': '25-35 LPA',
        'location': 'Bangalore',
        'skills': ['Python', 'DSA', 'System Design', 'Problem Solving'],
        'logo': 'fab fa-google'
    },
    {
        'title': 'Data Scientist',
        'company': 'Microsoft',
        'min_cgpa': 7.5,
        'min_skill': 80,
        'salary': '20-30 LPA',
        'location': 'Hyderabad',
        'skills': ['Python', 'Machine Learning', 'Statistics', 'SQL'],
        'logo': 'fab fa-microsoft'
    },
    {
        'title': 'Full Stack Developer',
        'company': 'Amazon',
        'min_cgpa': 7.0,
        'min_skill': 78,
        'salary': '18-25 LPA',
        'location': 'Chennai',
        'skills': ['React', 'Node.js', 'MongoDB', 'AWS'],
        'logo': 'fab fa-amazon'
    },
    {
        'title': 'Business Analyst',
        'company': 'Deloitte',
        'min_cgpa': 6.5,
        'min_skill': 70,
        'salary': '8-12 LPA',
        'location': 'Mumbai',
        'skills': ['Excel', 'SQL', 'Communication', 'Tableau'],
        'logo': 'fas fa-chart-line'
    },
    {
        'title': 'Software Developer',
        'company': 'Infosys',
        'min_cgpa': 6.0,
        'min_skill': 65,
        'salary': '6-10 LPA',
        'location': 'Pune',
        'skills': ['Java', 'Spring Boot', 'SQL'],
        'logo': 'fas fa-code'
    },
    {
        'title': 'Web Developer',
        'company': 'TCS',
        'min_cgpa': 6.0,
        'min_skill': 60,
        'salary': '5-8 LPA',
        'location': 'Multiple',
        'skills': ['HTML/CSS', 'JavaScript', 'React'],
        'logo': 'fas fa-globe'
    },
    {
        'title': 'Data Analyst',
        'company': 'Accenture',
        'min_cgpa': 6.5,
        'min_skill': 72,
        'salary': '7-11 LPA',
        'location': 'Bangalore',
        'skills': ['Python', 'Pandas', 'Power BI', 'Statistics'],
        'logo': 'fas fa-database'
    },
    {
        'title': 'DevOps Engineer',
        'company': 'IBM',
        'min_cgpa': 7.0,
        'min_skill': 75,
        'salary': '10-15 LPA',
        'location': 'Pune',
        'skills': ['Docker', 'Kubernetes', 'Jenkins', 'AWS'],
        'logo': 'fab fa-ibm'
    }
]

# ==================== AUTHENTICATION VIEWS ====================

def login_view(request):
    if request.user.is_authenticated:
        return redirect('about')
    return render(request, 'login.html')

def register_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        
        if password != confirm_password:
            messages.error(request, 'Passwords do not match!')
            return render(request, 'register.html')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists!')
            return render(request, 'register.html')
        
        if User.objects.filter(email=email).exists():
            messages.error(request, 'Email already registered!')
            return render(request, 'register.html')
        
        user = User.objects.create_user(username=username, email=email, password=password)
        user.save()
        messages.success(request, 'Account created successfully! Please login.')
        return redirect('login')
    
    return render(request, 'register.html')

# ==================== MAIN PAGE VIEWS ====================

@login_required(login_url='login')
def home(request):
    return render(request, 'home.html')

@login_required(login_url='login')
def about(request):
    return render(request, 'about.html')

@login_required(login_url='login')
def dashboard(request):
    # Get all predictions from database
    if MODELS_AVAILABLE:
        all_predictions = Prediction.objects.all()
        total_predictions = all_predictions.count()
        placed_count = all_predictions.filter(status='Placed').count()
        
        if total_predictions > 0:
            placement_rate = (placed_count / total_predictions) * 100
            avg_skill_score = all_predictions.aggregate(Avg('skill_score'))['skill_score__avg'] or 0
            avg_probability = all_predictions.aggregate(Avg('probability'))['probability__avg'] or 0
        else:
            placement_rate = 0
            avg_skill_score = 0
            avg_probability = 0
        
        recent_predictions = all_predictions.order_by('-timestamp')[:10]
        
        # Convert to list of dicts for template
        recent_list = []
        for pred in recent_predictions:
            recent_list.append({
                'timestamp': pred.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                'status': pred.status,
                'probability': pred.probability,
                'skill_score': pred.skill_score,
                'cgpa': pred.cgpa
            })
    else:
        # Fallback to session/history if database not available
        total_predictions = 0
        placed_count = 0
        placement_rate = 0
        avg_skill_score = 0
        avg_probability = 0
        recent_list = []
    
    stats = {
        'total_predictions': total_predictions,
        'placed_count': placed_count,
        'placement_rate': round(placement_rate, 1),
        'avg_skill_score': round(avg_skill_score, 1),
        'avg_probability': round(avg_probability, 1),
        'recent_predictions': recent_list
    }
    
    return render(request, 'dashboard.html', stats)

@login_required(login_url='login')
def profile(request):
    if MODELS_AVAILABLE:
        user_predictions = Prediction.objects.filter(user=request.user).order_by('-timestamp')
        total = user_predictions.count()
        placed = user_predictions.filter(status='Placed').count()
        placement_rate = (placed/total*100) if total > 0 else 0
        
        # Get statistics
        avg_skill = user_predictions.aggregate(Avg('skill_score'))['skill_score__avg'] or 0
        avg_prob = user_predictions.aggregate(Avg('probability'))['probability__avg'] or 0
        
        # Get latest prediction for job recommendations
        latest_pred = user_predictions.first()
        if latest_pred:
            job_recommendations = get_job_recommendations(latest_pred.cgpa, latest_pred.skill_score, latest_pred.course)
        else:
            job_recommendations = []
        
        context = {
            'predictions': user_predictions[:10],
            'total_predictions': total,
            'placement_rate': round(placement_rate, 1),
            'avg_skill_score': round(avg_skill, 1),
            'avg_probability': round(avg_prob, 1),
            'job_recommendations': job_recommendations,
            'has_predictions': total > 0
        }
    else:
        context = {
            'predictions': [],
            'total_predictions': 0,
            'placement_rate': 0,
            'avg_skill_score': 0,
            'avg_probability': 0,
            'job_recommendations': [],
            'has_predictions': False
        }
    
    return render(request, 'profile.html', context)

# ==================== JOB RECOMMENDATION FUNCTION ====================

def get_job_recommendations(cgpa, skill_score, course):
    """Get job recommendations based on student profile"""
    recommended = []
    
    for job in JOBS_DATABASE:
        if cgpa >= job['min_cgpa'] and skill_score >= job['min_skill']:
            job_copy = job.copy()
            # Calculate match percentage
            match = 0
            if cgpa >= job['min_cgpa'] + 1:
                match += 20
            elif cgpa >= job['min_cgpa']:
                match += 10
            
            if skill_score >= job['min_skill'] + 15:
                match += 20
            elif skill_score >= job['min_skill']:
                match += 10
            
            job_copy['match_percentage'] = min(95, 50 + match)
            recommended.append(job_copy)
    
    # Sort by match percentage
    recommended.sort(key=lambda x: x['match_percentage'], reverse=True)
    return recommended[:5]

# ==================== PREDICTION API ====================

def generate_recommendations(cgpa, aptitude, skill, internship, workex, probability):
    recommendations = []
    
    if cgpa < 6.0:
        recommendations.append("📚 Improve your CGPA - focus on core subjects and maintain consistency")
    elif cgpa < 7.0:
        recommendations.append("📖 Your CGPA is good, but aim for above 7.5 for better opportunities")
    elif cgpa >= 8.0:
        recommendations.append("🎓 Excellent CGPA! Highlight this in your resume")
    
    if aptitude < 60:
        recommendations.append("🧠 Practice aptitude tests daily - use platforms like Indiabix, PrepInsta")
    elif aptitude < 75:
        recommendations.append("📊 Your aptitude is decent, practice more complex problems and time management")
    else:
        recommendations.append("🎯 Great aptitude score! Focus on speed and accuracy")
    
    if skill < 70:
        recommendations.append("💻 Enhance technical skills - take online courses on Coursera/Udemy")
    elif skill < 85:
        recommendations.append("🚀 Good skills! Build projects to showcase your expertise")
    else:
        recommendations.append("⭐ Outstanding skills! Consider contributing to open source")
    
    if internship == 0:
        recommendations.append("🏢 Apply for internships - practical experience increases chances by 40%")
    elif internship < 2:
        recommendations.append("💼 Great start! Try to get one more internship")
    else:
        recommendations.append("🌟 Multiple internships will boost your resume significantly")
    
    if workex == 'No':
        recommendations.append("📝 Fresher? Focus on projects and certifications")
    else:
        recommendations.append("💪 Your work experience is valuable - quantify achievements")
    
    if probability < 40:
        recommendations.append("⚠️ Consider higher studies or skill development courses to improve chances")
        recommendations.append("🎯 Apply to startups and smaller companies first")
    elif probability < 70:
        recommendations.append("🎯 Moderate chances - apply to multiple companies and prepare well for interviews")
        recommendations.append("📢 Network on LinkedIn and attend placement workshops")
    else:
        recommendations.append("✨ Excellent chances! Focus on interview preparation and salary negotiation")
        recommendations.append("🏆 Target top companies and prepare for competitive exams")
    
    # Remove duplicates
    unique_recs = []
    for rec in recommendations:
        if rec not in unique_recs:
            unique_recs.append(rec)
    
    return unique_recs[:6]

@login_required(login_url='login')
@csrf_exempt
@require_http_methods(["POST"])
def predict_api(request):
    try:
        data = json.loads(request.body)
        
        # Extract all features
        gender = data.get('gender')
        hs_p = float(data.get('hs_p'))
        hs_b = data.get('hs_b')
        hs_12_p = float(data.get('12_p'))
        hs_12_b = data.get('12_b')
        hs_12_s = data.get('12_s')
        degree_p = float(data.get('degree_p'))
        degree_t = data.get('degree_t')
        workex = data.get('workex')
        etest_p = float(data.get('etest_p'))
        final_year_p = float(data.get('final_year_p'))
        course = data.get('course')
        course_specialization = data.get('course_specialization')
        skill_score = float(data.get('skill_score'))
        internship = int(data.get('internship'))
        
        # Encoding mappings
        gender_map = {'M': 0, 'F': 1}
        hs_b_map = {'Central': 0, 'Others': 1}
        hs_12_b_map = {'Central': 0, 'Others': 1}
        hs_12_s_map = {'Commerce': 0, 'Science': 1, 'Arts': 2}
        degree_t_map = {'Sci&Tech': 0, 'Comm&Mgmt': 1, 'Others': 2}
        workex_map = {'No': 0, 'Yes': 1}
        course_map = {'B.Tech': 0, 'BCA': 1, 'MCA': 2, 'MBA': 3, 'Pharmacy': 4}
        course_spec_map = {
            'ECE': 0, 'Web Dev': 1, 'Data Science': 2, 'App Dev': 3,
            'Finance': 4, 'Pharmacology': 5, 'Software': 6, 
            'Clinical Research': 7, 'HR': 8, 'CSE': 9, 'IT': 10
        }
        
        # Create input DataFrame
        input_data = pd.DataFrame([{
            'gender': gender_map.get(gender, 0),
            'hs p': hs_p,
            'hs b': hs_b_map.get(hs_b, 0),
            '12_p': hs_12_p,
            '12_b': hs_12_b_map.get(hs_12_b, 0),
            '12_s': hs_12_s_map.get(hs_12_s, 0),
            'degree_p': degree_p,
            'degree_t': degree_t_map.get(degree_t, 0),
            'workex': workex_map.get(workex, 0),
            'etest_p': etest_p,
            'final_year_p': final_year_p,
            'course': course_map.get(course, 0),
            'course_specialization': course_spec_map.get(course_specialization, 0),
            'skill_score': skill_score,
            'internship': internship
        }])
        
        # Scale and predict
        features_scaled = scaler.transform(input_data)
        prediction = model.predict(features_scaled)
        probability = model.predict_proba(features_scaled)[0][1]
        
        # Generate recommendations
        recommendations = generate_recommendations(
            degree_p, etest_p, skill_score, internship, workex, probability
        )
        
        result = {
            'success': True,
            'placement_status': 'Placed' if prediction[0] == 1 else 'Not Placed',
            'probability': round(probability * 100, 2),
            'recommendations': recommendations,
            'student_data': {
                'cgpa': degree_p,
                'aptitude_score': etest_p,
                'skill_score': skill_score,
                'internship': internship,
                'work_experience': workex,
                'final_year_percentage': final_year_p
            }
        }
        
        # Save to database if models are available
        if MODELS_AVAILABLE:
            try:
                Prediction.objects.create(
                    user=request.user,
                    status=result['placement_status'],
                    probability=result['probability'],
                    skill_score=int(skill_score),
                    cgpa=degree_p,
                    gender=gender,
                    hs_percentage=hs_p,
                    hs_board=hs_b,
                    twelfth_percentage=hs_12_p,
                    twelfth_board=hs_12_b,
                    twelfth_stream=hs_12_s,
                    degree_percentage=degree_p,
                    degree_type=degree_t,
                    work_experience=workex,
                    etest_percentage=etest_p,
                    final_year_percentage=final_year_p,
                    course=course,
                    course_specialization=course_specialization,
                    internship=internship
                )
                print(f"Prediction saved for user: {request.user.username}")
            except Exception as e:
                print(f"Error saving to database: {e}")
        
        return JsonResponse(result)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

# ==================== COMPARE VIEWS ====================

@login_required(login_url='login')
def compare(request):
    return render(request, 'compare.html')

@login_required(login_url='login')
@csrf_exempt
@require_http_methods(["POST"])
def compare_api(request):
    try:
        data = json.loads(request.body)
        student1 = data.get('student1', {})
        student2 = data.get('student2', {})
        
        score1 = calculate_success_score(student1)
        score2 = calculate_success_score(student2)
        breakdown1 = get_score_breakdown(student1)
        breakdown2 = get_score_breakdown(student2)
        
        return JsonResponse({
            'success': True,
            'student1_score': score1,
            'student2_score': score2,
            'student1_breakdown': breakdown1,
            'student2_breakdown': breakdown2,
            'winner': 'Student 1' if score1 > score2 else 'Student 2',
            'difference': abs(score1 - score2)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def calculate_success_score(data):
    try:
        score = 0
        score += float(data.get('degree_p', 0)) * 0.25
        score += float(data.get('final_year_p', 0)) * 0.15
        score += float(data.get('skill_score', 0)) * 0.25
        score += float(data.get('etest_p', 0)) * 0.05
        score += int(data.get('internship', 0)) * 3
        if data.get('workex') == 'Yes':
            score += 10
        return round(score, 2)
    except:
        return 0

def get_score_breakdown(data):
    try:
        academic = (float(data.get('degree_p', 0)) * 0.25 + float(data.get('final_year_p', 0)) * 0.15)
        skills = (float(data.get('skill_score', 0)) * 0.25 + float(data.get('etest_p', 0)) * 0.05)
        experience = (int(data.get('internship', 0)) * 3 + (10 if data.get('workex') == 'Yes' else 0))
        
        return {
            'academic': round(academic, 2),
            'skills': round(skills, 2),
            'experience': round(experience, 2),
            'total': round(academic + skills + experience, 2)
        }
    except:
        return {'academic': 0, 'skills': 0, 'experience': 0, 'total': 0}

# ==================== EXPORT FEATURES ====================

@login_required(login_url='login')
def download_report(request):
    """Generate and download PDF report"""
    if not REPORTLAB_AVAILABLE:
        return HttpResponse("ReportLab not installed. Run: pip install reportlab", status=400)
    
    try:
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        # Header
        p.setFillColor(HexColor('#1e3c72'))
        p.rect(0, height - 80, width, 80, fill=True)
        p.setFillColor(HexColor('#ffffff'))
        p.setFont("Helvetica-Bold", 24)
        p.drawString(50, height - 50, "Placement Prediction Report")
        
        # Student Info
        p.setFillColor(HexColor('#000000'))
        p.setFont("Helvetica", 12)
        p.drawString(50, height - 120, f"Student Name: {request.user.username}")
        p.drawString(50, height - 140, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        p.drawString(50, height - 160, f"Email: {request.user.email if request.user.email else 'Not provided'}")
        
        # Get last prediction from database
        if MODELS_AVAILABLE:
            last_pred = Prediction.objects.filter(user=request.user).order_by('-timestamp').first()
            if last_pred:
                y_position = height - 220
                p.setFont("Helvetica-Bold", 14)
                p.drawString(50, y_position, "Latest Prediction Result:")
                y_position -= 30
                
                p.setFont("Helvetica", 12)
                p.drawString(50, y_position, f"Status: {last_pred.status}")
                y_position -= 25
                p.drawString(50, y_position, f"Success Probability: {last_pred.probability}%")
                y_position -= 25
                p.drawString(50, y_position, f"Skill Score: {last_pred.skill_score}/100")
                y_position -= 25
                p.drawString(50, y_position, f"CGPA: {last_pred.cgpa}%")
                
                # Recommendations
                y_position -= 50
                p.setFont("Helvetica-Bold", 12)
                p.drawString(50, y_position, "Recommendations:")
                y_position -= 25
                p.setFont("Helvetica", 10)
                
                recs = generate_recommendations(
                    last_pred.cgpa, last_pred.etest_percentage, last_pred.skill_score,
                    last_pred.internship, last_pred.work_experience, last_pred.probability
                )
                
                for rec in recs[:4]:
                    if y_position > 50:
                        p.drawString(50, y_position, rec)
                        y_position -= 20
            else:
                p.drawString(50, height - 200, "No predictions found. Make a prediction first!")
        else:
            p.drawString(50, height - 200, "No predictions found. Make a prediction first!")
        
        # Footer
        p.setFont("Helvetica", 8)
        p.setFillColor(HexColor('#666666'))
        p.drawString(50, 50, "Generated by Placement Prediction System")
        p.drawString(50, 35, "© 2024 - AI-Powered Career Guidance")
        
        p.save()
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="placement_report.pdf"'
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating report: {str(e)}", status=400)

@login_required(login_url='login')
def export_excel(request):
    """Export prediction history to Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("OpenPyXL not installed. Run: pip install openpyxl", status=400)
    
    try:
        wb = openpyxl.Workbook()
        
        # Sheet 1: Predictions Data
        ws1 = wb.active
        ws1.title = "Predictions History"
        
        # Headers
        headers = ['Timestamp', 'Status', 'Probability (%)', 'Skill Score', 'CGPA (%)', 'Course', 'Internships']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e3c72", end_color="1e3c72", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Add data from database
        if MODELS_AVAILABLE:
            predictions = Prediction.objects.filter(user=request.user).order_by('-timestamp')
            for row, pred in enumerate(predictions, 2):
                ws1.cell(row=row, column=1, value=pred.timestamp.strftime("%Y-%m-%d %H:%M:%S"))
                ws1.cell(row=row, column=2, value=pred.status)
                ws1.cell(row=row, column=3, value=pred.probability)
                ws1.cell(row=row, column=4, value=pred.skill_score)
                ws1.cell(row=row, column=5, value=pred.cgpa)
                ws1.cell(row=row, column=6, value=pred.course)
                ws1.cell(row=row, column=7, value=pred.internship)
        else:
            ws1.cell(row=2, column=1, value="No predictions yet")
        
        # Adjust column widths
        for column in ws1.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="my_predictions.xlsx"'
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Error exporting Excel: {str(e)}", status=400)

@login_required(login_url='login')
@csrf_exempt
def share_result_api(request):
    """API for sharing results"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            result_data = data.get('result', {})
            
            share_text = f"🎓 Placement Prediction Result: {result_data.get('placement_status', 'N/A')} with {result_data.get('probability', 0)}% probability! Check your chances now!"
            
            return JsonResponse({
                'success': True,
                'share_text': share_text,
                'share_url': request.build_absolute_uri('/home/')
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})